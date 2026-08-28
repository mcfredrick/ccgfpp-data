#!/usr/bin/env python3
"""HISTORICAL / ONE-TIME MIGRATION TOOL. Airtable is the source of truth going
forward (see scripts/pull_airtable.py) -- this script is what originally
turned the 2022-2025 Google Sheet into tidy data before that migration, and
is kept for the record, not run as part of the regular pipeline anymore.

Parse the Food Pantry Plot xlsx export into tidy long-format records.

Input:  data/raw/*.xlsx  (native Sheets export via Drive download_file_content,
        exportMimeType=application/vnd.openxmlformats-officedocument.spreadsheetml.sheet)
Output: data/tidy/harvests.json, data/tidy/harvests.csv

Each tab = one year. Each row = one crop, one week, one quantity. `donated` is
True when the crop name carries a trailing "*" (sheet footnote: "donated from
other plots, gardeners"). `crop_raw` preserves the exact sheet spelling for
that year; `crop` is the canonical name after applying data/crop_mapping.json
(e.g. "Pepper"/"Peppers" both become "Peppers"), which is what charts and
totals should group by.

Sanity check: sums each crop's weekly values per year and compares against the
sheet's own "Season Total" column. This caught a real data-loss bug when we
were pulling via the natural-language markdown export instead of the native
xlsx (2022 Turnips/Winter Squash/Zucchini tail values were silently dropped) —
keep this check even though the xlsx path is exact, as a guard against future
sheet structure changes (inserted columns, renamed header, etc).
"""
import csv
import json
from pathlib import Path

import openpyxl

RAW_DIR = Path(__file__).parent.parent / "data" / "raw"
TIDY_DIR = Path(__file__).parent.parent / "data" / "tidy"
CROP_MAPPING_PATH = Path(__file__).parent.parent / "data" / "crop_mapping.json"


def load_crop_mapping():
    mapping = json.loads(CROP_MAPPING_PATH.read_text())
    mapping.pop("_comment", None)
    return mapping


def parse_workbook(path, crop_mapping):
    wb = openpyxl.load_workbook(path, data_only=True)
    records = []
    warnings = []
    for year_str in wb.sheetnames:
        try:
            year = int(year_str)
        except ValueError:
            continue
        ws = wb[year_str]
        rows = list(ws.iter_rows(values_only=True))
        header_idx = next(
            (i for i, r in enumerate(rows) if r and r[0] == "Crop" and r[1] == "Units"),
            None,
        )
        if header_idx is None:
            warnings.append(f"{year_str}: no 'Crop | Units | Season Total' header found, skipping tab")
            continue
        header = rows[header_idx]
        dates = header[3:]

        for row in rows[header_idx + 1:]:
            if not row or not row[0]:
                continue
            crop_raw = str(row[0]).strip()
            if crop_raw.lower().startswith("*donated") or crop_raw.lower().startswith("donated"):
                continue  # footnote row, not a crop
            unit = (row[1] or "").strip() if isinstance(row[1], str) else row[1]
            season_total = row[2]
            donated = crop_raw.endswith("*")
            crop_clean = crop_raw.rstrip("*").strip()
            crop_canonical = crop_mapping.get(crop_clean, crop_clean)

            running = 0
            for date_cell, qty_cell in zip(dates, row[3:]):
                if date_cell is None or qty_cell is None:
                    continue
                date_iso = date_cell.date().isoformat() if hasattr(date_cell, "date") else str(date_cell)
                records.append({
                    "year": year,
                    "crop": crop_canonical,
                    "crop_raw": crop_clean,
                    "unit": unit,
                    "date": date_iso,
                    "quantity": qty_cell,
                    "donated": donated,
                })
                running += qty_cell

            if season_total is not None and abs(running - season_total) > 0.01:
                warnings.append(
                    f"{year} {crop_clean}: weekly sum {running} != season total {season_total}"
                )
    return records, warnings


def main():
    xlsx_files = sorted(RAW_DIR.glob("*.xlsx"))
    if not xlsx_files:
        raise SystemExit("No .xlsx files found in data/raw/. Pull the sheet via Drive download_file_content first.")

    crop_mapping = load_crop_mapping()
    all_records = []
    all_warnings = []
    for path in xlsx_files:
        records, warnings = parse_workbook(path, crop_mapping)
        all_records.extend(records)
        all_warnings.extend(warnings)

    TIDY_DIR.mkdir(parents=True, exist_ok=True)
    (TIDY_DIR / "harvests.json").write_text(json.dumps(all_records, indent=2))

    with (TIDY_DIR / "harvests.csv").open("w", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=["year", "crop", "crop_raw", "unit", "date", "quantity", "donated"])
        writer.writeheader()
        writer.writerows(all_records)

    years = sorted(set(r["year"] for r in all_records))
    print(f"Parsed {len(all_records)} records across years: {years}")
    print(f"Distinct canonical crops: {len(set(r['crop'] for r in all_records))}")
    if all_warnings:
        print(f"\n{len(all_warnings)} reconciliation warnings (weekly sum vs Season Total column):")
        for w in all_warnings:
            print(f"  - {w}")
    else:
        print("\nAll crop-year weekly sums reconcile with the sheet's Season Total column.")


if __name__ == "__main__":
    main()
